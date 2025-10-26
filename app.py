from flask import Flask, render_template, request, jsonify, send_file
import sqlite3, os, io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), "database", "bearings.db")


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/bearings")
def api_bearings():
    conn = get_db_connection()
    bearings = conn.execute("SELECT * FROM bearings ORDER BY id DESC").fetchall()
    conn.close()
    return jsonify([dict(row) for row in bearings])


@app.route("/generate", methods=["POST"])
def generate():
    data = request.json

    # Apply uppercase and SAP rules
    bearing_type = data["bearing_type"].upper().strip()
    bearing_sub_type = data["bearing_sub_type"].upper().strip()
    bearing_number = data["bearing_number"].upper().strip()
    seals_shields = data["seals_shields"].upper().strip()
    other_suffixes = data["other_suffixes"].upper().strip()
    make_application = data["make_application"].upper().strip()

    spec = f"{bearing_type} {bearing_sub_type} {bearing_number} {seals_shields} {other_suffixes}".strip()[
        :40
    ]
    full_description = (
        f"{bearing_type} ({bearing_sub_type} SERIES) NUMBER: {bearing_number}, "
        f"SEALS/SHIELDS: {seals_shields}, SUFFIXES: {other_suffixes}, "
        f"APPLICATION: {make_application}"
    )[:255]

    conn = get_db_connection()
    conn.execute(
        """
        INSERT INTO bearings 
        (bearing_type, bearing_sub_type, bearing_number, seals_shields, other_suffixes, make_application, generated_specification)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """,
        (
            bearing_type,
            bearing_sub_type,
            bearing_number,
            seals_shields,
            other_suffixes,
            make_application,
            spec,
        ),
    )
    conn.commit()
    conn.close()

    # Generate Excel and send as download
    return generate_excel_download()


def generate_excel_download():
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM bearings ORDER BY id ASC").fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bearings"

    headers = [
        "ID",
        "Type",
        "Sub-Type",
        "Number",
        "Seals",
        "Suffixes",
        "Application",
        "Generated Spec",
    ]
    ws.append(headers)

    # Color fills
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )  # ✅ Correct
    yellow_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )  # ⚠️ Review
    red_fill = PatternFill(
        start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
    )  # ❌ Wrong

    for row in rows:
        generated_spec = row["generated_specification"].upper()

        # Determine classification color
        if any(
            keyword in generated_spec
            for keyword in ["DEEP GROOVE", "CYLINDRICAL", "SPHERICAL"]
        ):
            fill = green_fill
        elif any(keyword in generated_spec for keyword in ["NEEDLE", "903"]):
            fill = yellow_fill
        else:
            fill = red_fill

        excel_row = [
            row["id"],
            row["bearing_type"],
            row["bearing_sub_type"],
            row["bearing_number"],
            row["seals_shields"],
            row["other_suffixes"],
            row["make_application"],
            generated_spec,
        ]
        ws.append(excel_row)

        # Apply color to the last added row
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = Font(name="Calibri", size=11)

    # Auto column width
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="bearing_specifications.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
